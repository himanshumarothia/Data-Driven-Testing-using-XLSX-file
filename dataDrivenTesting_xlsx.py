from selenium import webdriver
from selenium.webdriver.common.by import By 
from openpyxl import load_workbook
import time 

#load the excelsheet 
workbook = load_workbook('Book1.xlsx')

#selecting Active sheet 
sheet = workbook.active

driver = webdriver.Firefox()
driver.maximize_window()

time.sleep(1)

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True) :
    username = row[0]
    password = row[1]

    driver.get("https://www.saucedemo.com/")
    time.sleep(1)
    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    time.sleep(1)
    driver.find_element(By.XPATH,"//input[@id='login-button']").click()


    time.sleep(1)
    driver.find_element(By.XPATH,"//button[@id='react-burger-menu-btn']").click()

    time.sleep(1)
    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()

    time.sleep(1)
    

driver.quit()
